from openpyxl import load_workbook
from all_paths import *

excel = load_workbook(r"C:\Users\User\Desktop\Timesheet\2023\July 23\Invoice_July23.xlsx")
sheet = excel['Sheet2']

Mar_Attendance = Attendance()
Mar_Attendance.login()

for i in range(2, sheet.max_row+1):
    Mar_Attendance.add_attendance() 
    Mar_Attendance.month_dropdown()
    Mar_Attendance.year_dropdown()
    Mar_Attendance.worked_days(value=sheet.cell(row=i, column=1).value)
    Mar_Attendance.leaves(value=sheet.cell(row=i, column=2).value)
    Mar_Attendance.project_dropdown(value=sheet.cell(row=i, column=3).value)
    Mar_Attendance.employee_id(value=sheet.cell(row=i, column=4).value)
    #Mar_Attendance.verification()
    
    if Mar_Attendance.verification()==True:
        sheet.cell(row=i, column=5).value="Done"
        excel.save(r"C:\Users\User\Desktop\Timesheet\2023\July 23\Invoice_July23.xlsx")
    print('-' * 60)

print("Attendance is added successfully for all the employees ")

